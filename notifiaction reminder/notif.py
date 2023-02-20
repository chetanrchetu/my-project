import datetime
from email import message
from socket import timeout 
from stat import S_ENFMT
import openpyxl as ox
from pathlib import Path
from plyer import *
import time

xls=Path(Path.cwd(),"Book1.xlsx")
w=ox.load_workbook(xls)
sheet=w.active
#starts from 2
size=sheet.max_row
lis=[]
for i in range(2,size+1):
    si=sheet.cell(row=i,column=1).value
    s=int(si[0])
    lis.append([s,si[0+5]])



 
for i in range(0,len(lis)):
    s=int(lis[i][0])
   
    e=int(lis[i][1])
    
    
    h=datetime.datetime.now().hour
    m=datetime.datetime.now().minute
    
    diff=e-s
    tt=sheet.cell(row=i+2,column=2).value
    mm=sheet.cell(row=i+2,column=3).value
    print((diff*3600)-(m*60))
    print(s,h%12,e)
    s1=str(s)+"00"
    e1=str(e)+"00"
    
    vaa=""
    if m<=10:
        vaa+="0"+str(m)
    else:
        vaa=str(m)

        
    h1=str(h%12)+vaa
    
    if(int(s1)<=int(h1)<=int(e1)):
        notification.notify(
            title=str(tt),
            message=str(mm),
            timeout=10
        )

        time.sleep((diff*3600)-(m*60))

    
    
    
        



